import pandas as pd
import numpy as np
import random
import math
import html
import os
import tempfile
import webbrowser
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Pane


def generate_color():
    return tuple(random.randint(50, 200) for _ in range(3))


def get_contrasting_color(rgb):
    r, g, b = rgb
    brightness = (r * 299 + g * 587 + b * 114) / 1000
    return (0, 0, 0) if brightness > 125 else (255, 255, 255)


def ansi_style(text, bg=None, fg=None):
    seq = ""
    if bg:
        seq += f"\033[48;2;{bg[0]};{bg[1]};{bg[2]}m"
    if fg:
        seq += f"\033[38;2;{fg[0]};{fg[1]};{fg[2]}m"
    return f"{seq}{text}\033[0m"


def truncate_text(text, max_length):
    if len(text) <= max_length:
        return text
    return text[:max_length - 3] + "..."


def handle_nan_value(value, nan_placeholder="N/A"):
    """Handle NaN or None values with a placeholder"""
    if isinstance(value, list):
        return repr(value)
    if pd.isna(value):
        return nan_placeholder
    return str(value)


def draw_table_border(widths, style="ascii", is_header=False):
    """Draw table borders using ASCII or Unicode characters"""
    if style == "none":
        return ""

    # Character sets for different border styles
    border_chars = {
        "ascii": {
            "top_left": "+", "top_right": "+", "bottom_left": "+", "bottom_right": "+",
            "top_middle": "+", "bottom_middle": "+", "left_middle": "+", "right_middle": "+",
            "middle": "+", "horizontal": "-", "vertical": "|"
        },
        "unicode": {
            "top_left": "┌", "top_right": "┐", "bottom_left": "└", "bottom_right": "┘",
            "top_middle": "┬", "bottom_middle": "┴", "left_middle": "├", "right_middle": "┤",
            "middle": "┼", "horizontal": "─", "vertical": "│"
        },
        "unicode_heavy": {
            "top_left": "┏", "top_right": "┓", "bottom_left": "┗", "bottom_right": "┛",
            "top_middle": "┳", "bottom_middle": "┻", "left_middle": "┣", "right_middle": "┫",
            "middle": "╋", "horizontal": "━", "vertical": "┃"
        }
    }

    chars = border_chars.get(style, border_chars["ascii"])

    # Create the appropriate border line based on position
    if is_header:
        middle_char = chars["middle"]
        left_char = chars["left_middle"]
        right_char = chars["right_middle"]
    else:
        if is_header is None:  # Top border
            middle_char = chars["top_middle"]
            left_char = chars["top_left"]
            right_char = chars["top_right"]
        else:  # Bottom border
            middle_char = chars["bottom_middle"]
            left_char = chars["bottom_left"]
            right_char = chars["bottom_right"]

    result = left_char
    for i, width in enumerate(widths):
        result += chars["horizontal"] * (width + 2)  # +2 for padding spaces
        result += middle_char if i < len(widths) - 1 else right_char

    return result


def find_column_with_fewest_unique_values(df):
    """Find the column with the fewest unique values to use as default grouping column"""
    if df.empty or len(df.columns) == 0:
        return None

    # Calculate the number of unique values in each column
    unique_counts = {col: df[col].nunique() for col in df.columns}

    # Find the column with the minimum number of unique values
    min_col = min(unique_counts, key=unique_counts.get)

    # Only use columns with at least 2 unique values and less than half the number of rows
    good_cols = {col: count for col, count in unique_counts.items()
                 if count >= 2 and count < len(df) / 2}

    if good_cols:
        return min(good_cols, key=good_cols.get)

    return min_col


def rgb_to_hex(rgb):
    """Convert RGB tuple to hex color string"""
    return f"#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"

#Principal
def generate_excel_table(
        df_display,
        group_col,
        special_col,
        col_widths,
        group_colors,
        special_colors,
        style_mode,
        persist_special_color,
        nan_placeholder,
        border_style,
        max_height=None,
        synchronized_panels=False
    ):

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Styled Table"
    
    # Define border styles based on the input parameter
    border_styles = {
        "none": None,
        "ascii": Side(style='thin', color='DDDDDD'),
        "unicode": Side(style='thin', color='DDDDDD'),
        "unicode_heavy": Side(style='medium', color='000000')
    }
    
    border_side = border_styles.get(border_style)
    
    # Add headers to the first row
    for col_idx, col_name in enumerate(df_display.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        
        # Style the header cells
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Apply border if specified
        if border_side:
            cell.border = Border(top=border_side, left=border_side, right=border_side, bottom=border_side)
        
        # Set alignment
        cell.alignment = Alignment(horizontal='left')
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Add data rows
    for row_idx, (_, row_data) in enumerate(df_display.iterrows(), 2):
        # Get group value color
        group_val = row_data[group_col]
        row_color = group_colors[group_val]
        text_color = get_contrasting_color(row_color)
        
        # Convert to hex for Excel
        row_color_hex = rgb_to_hex(row_color)[1:]
        text_color_hex = rgb_to_hex(text_color)[1:]
        
        # Process each column in the row
        for col_idx, col_name in enumerate(df_display.columns, 1):
            # Handle NaN values
            val = handle_nan_value(row_data[col_name], nan_placeholder)
            
            # Add the value to the cell
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            
            # Apply base styling (group-based)
            if style_mode == "background":
                cell.fill = PatternFill(start_color=row_color_hex.upper(), end_color=row_color_hex.upper(), fill_type="solid")
                cell.font = Font(color=text_color_hex.upper())
            else:  # text mode
                cell.font = Font(color=row_color_hex.upper())
            
            # Apply special coloring if applicable
            if special_col is not None:
                special_val = row_data[special_col]
                if pd.isna(special_val):
                    special_val = np.nan
                
                special_color = special_colors[special_val]
                special_text_color = get_contrasting_color(special_color)
                
                # Convert to hex
                special_color_hex = rgb_to_hex(special_color)[1:]
                special_text_color_hex = rgb_to_hex(special_text_color)[1:]
                
                is_special = (col_name == special_col)
                apply_special = is_special or (
                    persist_special_color and 
                    col_idx > df_display.columns.get_loc(special_col) + 1
                )
                
                if apply_special:
                    if style_mode == "text":
                        cell.font = Font(color=special_color_hex.upper())
                    else:  # background
                        cell.fill = PatternFill(start_color=special_color_hex.upper(), end_color=special_color_hex.upper(), fill_type="solid")
                        cell.font = Font(color=special_text_color_hex.upper())
            
            # Apply border if specified
            if border_side:
                cell.border = Border(top=border_side, left=border_side, right=border_side, bottom=border_side)
            
            # Set alignment
            cell.alignment = Alignment(horizontal='left')
    
    # Auto-adjust column widths based on content and specified widths
    for col_idx, col_name in enumerate(df_display.columns, 1):
        excel_width = col_widths[col_name]
        ws.column_dimensions[get_column_letter(col_idx)].width = excel_width
    ws.auto_filter.ref = ws.dimensions
    # split_column = None
    # if synchronized_panels:
    #     # Calculate approximately how many columns would fit on screen
    #     # Assuming an average screen width of around 1200-1400 pixels
    #     # and accounting for potential scrollbars and margins
        
    #     # Calculate total width of all columns
    #     total_width = sum(col_widths.values())
        
    #     # Estimate visible width (this could be passed as a parameter)
    #     estimated_visible_width = 250  # Adjust this based on your typical display
        
    #     # Calculate how many columns would be visible
    #     cumulative_width = 0
    #     visible_columns = 0
        
    #     for col_name in df_display.columns:
    #         cumulative_width += col_widths[col_name]
    #         visible_columns += 1
    #         if cumulative_width > estimated_visible_width:
    #             break
    #     # Set the split at the middle of visible columns
    #     if visible_columns > 1:
    #         split_column = visible_columns // 2
    #         split_column_letter = get_column_letter(split_column + 1)
            
    #         # # Apply the split view
    #         # sheet_view = SheetView()
    #         # ws.sheet_view = sheet_view
    #         # pane = Pane(
    #         #     xSplit=6,
    #         #     ySplit=0,
    #         #     topLeftCell="K1", 
    #         #     activePane="topLeft",
    #         #     state="split"
    #         # )
    #         # ws.sheet_view.pane = pane
    #         # ws.sheet_view.showGridLines = True
    #     else:
    #         # If only one column would be visible, just use the default freeze panes
    #         ws.freeze_panes = 'B2'

    return wb


def generate_html_table(
        df_display,
        group_col,
        special_col,
        col_widths,
        group_colors,
        special_colors,
        style_mode,
        persist_special_color,
        nan_placeholder,
        border_style,
        max_height=None,
        synchronized_panels=False  # New parameter
    ):
    """Generate HTML table with similar styling to the terminal output"""
    
    # Define CSS styles for the table based on border style
    border_styles = {
        "none": "border: none;",
        "ascii": "border: 1px solid #ddd;",
        "unicode": "border: 1px solid #ddd;",
        "unicode_heavy": "border: 2px solid #000;"
    }
    
    # Set up the table style
    table_border = border_styles.get(border_style, "border: none;")
    
    # Create the HTML table content (without container)
    table_html = f"""
      <table style="{table_border} border-collapse: collapse; width: 100%; font-family: monospace;">
        <thead>
          <tr>
    """
    
    # Add header cells
    for col in df_display.columns:
        col_str = str(col)
        if len(col_str) > col_widths[col]:
            col_str = truncate_text(col_str, col_widths[col])
        
        # Make the header always visible
        header_style = "padding: 8px; text-align: left; font-weight: bold; position: sticky; top: 0; background-color: #f8f9fa; z-index: 1;"
        if border_style != "none":
            header_style += " border: 1px solid #ddd;"
            
        table_html += f'<th style="{header_style}">{html.escape(col_str)}</th>'
    
    table_html += """
          </tr>
        </thead>
        <tbody>
    """
    
    # Add data rows
    for _, row in df_display.iterrows():
        table_html += "<tr>"
        
        # Get group value color
        group_val = row[group_col]
        row_color = group_colors[group_val]
        text_color = get_contrasting_color(row_color)
        
        # Convert to hex for HTML
        row_color_hex = rgb_to_hex(row_color)
        text_color_hex = rgb_to_hex(text_color)
        
        # Process each column in the row
        for col in df_display.columns:
            # Handle NaN values
            val = handle_nan_value(row[col], nan_placeholder)
            original_val = val  # Save the original complete value
            
            # Truncate long values for display
            truncated = False
            if len(val) > col_widths[col]:
                display_val = truncate_text(val, col_widths[col])
                truncated = True
            else:
                display_val = val
            
            # Default cell styling based on group
            cell_style = "padding: 8px; cursor: pointer;"
            if border_style != "none":
                cell_style += " border: 1px solid #ddd;"
                
            # Apply base styling (group-based)
            if style_mode == "background":
                cell_style += f" background-color: {row_color_hex}; color: {text_color_hex};"
            else:  # text mode
                cell_style += f" color: {row_color_hex};"
                
            # Apply special coloring if applicable
            if special_col is not None:
                special_val = row[special_col]
                if pd.isna(special_val):
                    special_val = np.nan
                    
                special_color = special_colors[special_val]
                special_text_color = get_contrasting_color(special_color)
                
                # Convert to hex
                special_color_hex = rgb_to_hex(special_color)
                special_text_color_hex = rgb_to_hex(special_text_color)
                
                is_special = (col == special_col)
                apply_special = is_special or (
                    persist_special_color and 
                    col in df_display.columns[df_display.columns.get_loc(special_col) + 1:]
                )
                
                if apply_special:
                    if style_mode == "text":
                        cell_style = cell_style.replace(f"color: {row_color_hex};", f"color: {special_color_hex};")
                    else:  # background
                        cell_style = cell_style.replace(
                            f"background-color: {row_color_hex}; color: {text_color_hex};",
                            f"background-color: {special_color_hex}; color: {special_text_color_hex};"
                        )
            
            # Add style to truncate text
            cell_style += " white-space: nowrap; overflow: hidden; text-overflow: ellipsis;"
            
            # Add cell with styled content and data attribute for the complete text
            if truncated:
                table_html += f'<td style="{cell_style}" class="expandable-cell" data-full-text="{html.escape(original_val)}">{html.escape(display_val)}</td>'
            else:
                table_html += f'<td style="{cell_style}">{html.escape(display_val)}</td>'
            
        table_html += "</tr>"
    
    # Close the table
    table_html += """
        </tbody>
      </table>
    """
    # Set up container style with max-height if specified
    container_style = "width: 100%; overflow-x: auto; overflow-y: auto;"
    if max_height:
        container_style += f" max-height: {max_height}px;"

    # Create the complete HTML content based on whether synchronized panels are requested
    if synchronized_panels:
        html_output = f"""
        <div style="width: 100%;">
            <div style="display: flex; flex-direction: column;">
                <h3>Synchronized Panels (continuous view)</h3>
                <div style="display: flex; width: 100%; border: 1px solid #ccc; margin-bottom: 10px;">
                    <div id="leftPanel" class="sync-panel" style="flex: 1; overflow: auto; height: {max_height}px; border-right: 2px dashed #666;">
                        {table_html}
                    </div>
                    <div id="rightPanel" class="sync-panel" style="flex: 1; overflow: auto; height: {max_height}px;">
                        {table_html}
                    </div>
                </div>
            </div>
            <p>{html.escape(group_info)}</p>
        </div>
        
        <script>
            document.addEventListener('DOMContentLoaded', function() {{
                // Expandable cell functionality
                const cells = document.querySelectorAll('.expandable-cell');
                cells.forEach(cell => {{
                    // Guardar el texto original de la celda si no está almacenado
                    if (!cell.hasAttribute('data-original-text')) {{
                        cell.setAttribute('data-original-text', cell.textContent);
                    }}

                    cell.addEventListener('click', function(e) {{
                        e.stopPropagation(); // Prevent event from reaching the row
                        
                        const fullText = this.getAttribute('data-full-text');
                        
                        // If already expanded, contract it
                        if (this.classList.contains('expanded')) {{
                            this.classList.remove('expanded');
                            this.textContent = this.getAttribute('data-original-text'); // Restoring original text
                        }} 
                        // If not expanded, expand it
                        else {{
                            // Contract all other cells first
                            document.querySelectorAll('.expandable-cell.expanded').forEach(expandedCell => {{
                                expandedCell.classList.remove('expanded');
                                expandedCell.textContent = expandedCell.getAttribute('data-original-text'); // Restoring original text
                            }});
                            
                            this.classList.add('expanded');
                            this.textContent = fullText; // Show full text
                        }}
                    }});
                }});

                // Selección de fila al hacer doble clic - VERSIÓN CORREGIDA
                function toggleRowSelection(row) {{
                    const rowsLeft = document.querySelectorAll('#leftPanel tbody tr');
                    const rowsRight = document.querySelectorAll('#rightPanel tbody tr');
                    
                    // Eliminar la clase "selected-row" de todas las filas en ambos paneles
                    rowsLeft.forEach(r => r.classList.remove('selected-row'));
                    rowsRight.forEach(r => r.classList.remove('selected-row'));
                    
                    // Determinar en qué panel está la fila seleccionada
                    const isLeftPanel = row.closest('#leftPanel') !== null;
                    
                    // Determinar el índice de la fila seleccionada
                    const rowIndex = Array.from(isLeftPanel ? rowsLeft : rowsRight).indexOf(row);
                    
                    // Agregar la clase "selected-row" a la fila seleccionada en ambos paneles
                    if (rowIndex >= 0) {{
                        // Seleccionar la fila en el panel izquierdo
                        if (rowIndex < rowsLeft.length) {{
                            rowsLeft[rowIndex].classList.add('selected-row');
                        }}
                        
                        // Seleccionar la fila en el panel derecho
                        if (rowIndex < rowsRight.length) {{
                            rowsRight[rowIndex].classList.add('selected-row');
                        }}
                    }}
                }}

                // Agregar el evento de doble clic para las filas en ambos paneles
                const rowsLeft = document.querySelectorAll('#leftPanel tbody tr');
                const rowsRight = document.querySelectorAll('#rightPanel tbody tr');

                // Aplicar el evento de doble clic en el panel izquierdo
                rowsLeft.forEach(row => {{
                    row.addEventListener('dblclick', function() {{
                        toggleRowSelection(this); // Selecciona la fila al hacer doble clic
                    }});
                }});

                // Aplicar el evento de doble clic en el panel derecho
                rowsRight.forEach(row => {{
                    row.addEventListener('dblclick', function() {{
                        toggleRowSelection(this); // Selecciona la fila al hacer doble clic
                    }});
                }});

                // Sincronizar el desplazamiento vertical entre los paneles
                const leftPanel = document.getElementById('leftPanel');
                const rightPanel = document.getElementById('rightPanel');
                
                // Calcular el ancho en píxeles y ajustar la posición del panel derecho
                const leftTable = leftPanel.querySelector('table');
                const headerCells = leftTable.querySelectorAll('th');

                // Establecer la posición inicial del panel derecho para continuar desde donde termina el panel izquierdo
                function updateRightPanelPosition() {{
                    const leftPanelWidth = leftPanel.clientWidth;
                    const visibleWidth = leftPanel.getBoundingClientRect().width;
                    const tableRect = leftTable.getBoundingClientRect();
                    
                    // Calculamos la cantidad de píxeles visibles de la tabla en el panel izquierdo
                    const visibleTableWidth = Math.min(tableRect.width, visibleWidth);
                    
                    // Obtenemos la posición exacta en píxeles donde debería comenzar el panel derecho
                    let scrollPosition = visibleTableWidth;
                    
                    // Ajuste para crear un pequeño solapamiento visual para una mejor transición (antes 20)
                    //scrollPosition -= 20;
                    
                    if (scrollPosition > 0) {{
                        rightPanel.scrollLeft = scrollPosition;
                    }}
                }}

                // Actualizar la posición cuando cambie el tamaño de la ventana
                window.addEventListener('resize', updateRightPanelPosition);

                // Actualizar la posición inicialmente
                updateRightPanelPosition();
                
                // Sincronizar desplazamiento vertical
                leftPanel.addEventListener('scroll', function() {{
                    rightPanel.scrollTop = leftPanel.scrollTop;
                }});
                
                rightPanel.addEventListener('scroll', function() {{
                    leftPanel.scrollTop = rightPanel.scrollTop;
                }});
            }});
        </script>

        <style>
            .expandable-cell.expanded {{
                white-space: normal !important;
                word-wrap: break-word !important;
                position: relative;
                z-index: 10;
                min-width: 200px;
                box-shadow: 0 0 10px rgba(0,0,0,0.1);
            }}
            
            .selected-row {{
                background-color: #ff0000 !important;
                outline: 2px solid #ff0000;
            }}
            
            /* Adding a visual indicator for the panel boundary */
            #leftPanel::after {{
                content: '';
                position: absolute;
                right: 0;
                top: 0;
                height: 100%;
                width: 2px;
                background-color: #666;
                box-shadow: 0 0 5px rgba(0,0,0,0.3);
            }}
        </style>
        """
    else:
        # Original single panel output
        html_output = f"""
        <div style="{container_style}">
            {table_html}
        </div>
        
        <p>{html.escape(group_info)}</p>
        
        <script>
            document.addEventListener('DOMContentLoaded', function() {{
                // Expandable cell functionality
                const cells = document.querySelectorAll('.expandable-cell');
                cells.forEach(cell => {{
                    // Guardar el texto original de la celda si no está almacenado
                    if (!cell.hasAttribute('data-original-text')) {{
                        cell.setAttribute('data-original-text', cell.textContent);
                    }}
                    cell.addEventListener('click', function(e) {{
                        e.stopPropagation(); // Prevent event from reaching the row
                        
                        const fullText = this.getAttribute('data-full-text');
                        
                        // If already expanded, contract it
                        if (this.classList.contains('expanded')) {{
                            this.classList.remove('expanded');
                            this.textContent = this.getAttribute('data-original-text'); // Restore truncated text
                        }} 
                        // If not expanded, expand it
                        else {{
                            // Contract all other cells first
                            document.querySelectorAll('.expandable-cell.expanded').forEach(expandedCell => {{
                                expandedCell.classList.remove('expanded');
                                expandedCell.textContent = expandedCell.getAttribute('data-original-text'); // Restoring original text
                        }});
                            
                            this.classList.add('expanded');
                            this.textContent = fullText; // Show full text
                        }}
                }});
                }});
                
                // Select row on double-click
                const rows = document.querySelectorAll('tbody tr');
                rows.forEach(row => {{
                    row.addEventListener('dblclick', function() {{
                        // Remove selected-row class from all rows
                        rows.forEach(r => r.classList.remove('selected-row'));
                        // Add selected-row class to current row
                        this.classList.add('selected-row');
                    }});
                }});
            }});
        </script>
        
        <style>
            .expandable-cell.expanded {{
                white-space: normal !important;
                word-wrap: break-word !important;
                position: relative;
                z-index: 10;
                min-width: 200px;
                box-shadow: 0 0 10px rgba(0,0,0,0.1);
            }}
            
            .selected-row {{
                background-color: #ff0000 !important;
                outline: 2px solid #ff0000;
            }}
        </style>
        """
    
    return html_output


def print_colored_df(
        df,
        group_col=None,
        special_col=None,
        style_mode="background",
        persist_special_color=False,
        max_column_width=30,
        separator="  ",
        border_style="none",
        nan_placeholder="N/A",
        page_size=None,
        current_page=1,
        group_pagination=None,
        output_format="terminal",
        html_max_height=1000,
        synchronized_panels=False,
        output_name = "styled_table"
):
    """
    Display a DataFrame with colored formatting and various display options.

    Parameters:
    -----------
    df : pandas.DataFrame
        The DataFrame to display
    group_col : str, optional
        Column name used for grouping rows by color. If None, the column with fewest unique values is used
    special_col : str, optional
        Column name for special highlighting. If None, no special highlighting is applied
    style_mode : str, default="background"
        "background" for colored backgrounds, "text" for colored text
    persist_special_color : bool, default=True
        Whether to apply special color to columns after special_col
    max_column_width : int, default=20
        Maximum width for any column
    separator : str, default="  "
        Separator string between columns
    border_style : str, default="none"
        Border style: "none", "ascii", "unicode", or "unicode_heavy"
    nan_placeholder : str, default="N/A"
        String to display for NaN values
    page_size : int, optional
        Number of rows per page (None for all rows)
    current_page : int, default=1
        Current page to display when using pagination
    group_pagination : str or list, optional
        If provided, pagination will be by group(s) instead of row numbers:
        - If str: Show data for the specified group value
        - If list: Show data for the specified group values
        - If None: Pagination based on page_size and current_page
    output_format : str, default="terminal"
        Output format: "terminal" for console output or "html" for HTML output
    html_max_height : int, default=500
        Maximum height in pixels for the HTML table container (enables vertical scrolling)
    synchronized_panels : bool, default=False
        If True and output_format is "html", displays the table in two synchronized panels
        with independent horizontal scrolling and synchronized vertical scrolling
    """
    global group_info  # Make this accessible to the HTML generator function

    # Make a copy of the dataframe to avoid modifying the original
    df_full = df.copy()

    # Auto-select group_col if not specified
    if group_col is None:
        group_col = find_column_with_fewest_unique_values(df_full)
        if group_col:
            print(f"Auto-selected '{group_col}' as group column (fewest unique values: {df_full[group_col].nunique()})")
        else:
            print(f"Warning: Couldn't find a suitable grouping column. Using index instead.")
            df_full.reset_index(inplace=True)
            group_col = 'index'

    # Handle missing group column
    if group_col not in df_full.columns:
        print(f"Warning: group_col '{group_col}' not found in DataFrame. Using index instead.")
        df_full.reset_index(inplace=True)
        group_col = 'index'

    # Handle special_col
    if special_col is not None and special_col not in df_full.columns:
        print(f"Warning: special_col '{special_col}' not found in DataFrame. Not applying special highlighting.")
        special_col = None

    # Process group pagination if specified
    if group_pagination is not None:
        # Get all unique group values for reference
        all_groups = sorted(df_full[group_col].unique())

        if isinstance(group_pagination, (str, int, float)):
            # Single group specified
            group_pagination = [group_pagination]

        # Filter dataframe to only include specified groups
        df_display = df_full[df_full[group_col].isin(group_pagination)].copy()

        if len(df_display) == 0:
            print(f"No data found for group(s): {group_pagination}")
            print(f"Available groups: {all_groups}")
            return

        # Show info about which groups are being displayed
        total_groups = len(all_groups)
        current_groups = group_pagination

        group_info = f"Showing {len(current_groups)} of {total_groups} groups: {current_groups}"

    else:
        # Standard row-based pagination
        total_rows = len(df_full)

        if page_size is not None:
            total_pages = math.ceil(total_rows / page_size)
            current_page = max(1, min(current_page, total_pages))
            start_idx = (current_page - 1) * page_size
            end_idx = min(start_idx + page_size, total_rows)
            df_display = df_full.iloc[start_idx:end_idx].copy()

            # Pagination info for row-based pagination
            group_info = f"Page {current_page} of {total_pages} | Rows {start_idx + 1}-{end_idx} of {total_rows}"
        else:
            df_display = df_full.copy()
            group_info = f"Showing all {len(df_full)} rows"

    # Calculate column widths with maximum limit
    col_widths = {}
    for col in df_display.columns:
        header_len = len(str(col))
        # Handle NaN values when calculating width
        max_data_len = df_display[col].apply(lambda x: len(handle_nan_value(x, nan_placeholder))).max()
        col_widths[col] = min(max(header_len, max_data_len)*1.7, max_column_width) #1.7 

    # Widths list for border drawing
    widths_list = [col_widths[col] for col in df_display.columns]

    # Generate colors for groups
    group_colors = {g: generate_color() for g in df_full[group_col].unique()}

    # Generate colors for special values only if special_col is specified
    special_colors = {}
    if special_col is not None:
        special_colors = {s: generate_color() for s in df_full[special_col].dropna().unique()}
        # Add a default color for NaN values in special column
        special_colors[np.nan] = generate_color()

    # Handle excel output mode
    if output_format == "excel":
        excel_table = generate_excel_table(
            df_display,
            group_col,
            special_col,
            col_widths,
            group_colors,
            special_colors,
            style_mode,
            persist_special_color,
            nan_placeholder,
            border_style,
            html_max_height,
            synchronized_panels
        )

        root_name, ext_file = os.path.splitext(output_name)
        if ext_file.lower() != '.xlsx':
            output_name = root_name + '.xlsx'
        while True: 
            try: 
                excel_table.save(output_name)
                print(f"Excel file saved to: {os.path.abspath(output_name)}")
                break 
            except: 
                input(f"Cierra {os.path.abspath(output_name)} y Enter para reintentar...")
        if os.name == 'nt':
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True
            workbook = excel.Workbooks.Open(os.path.abspath(output_name))
            if synchronized_panels:
                active_window = workbook.Parent.ActiveWindow
                visible_range = active_window.VisibleRange
                first_col = visible_range.Column
                last_col = visible_range.Columns(visible_range.Columns.Count).Column
                active_window = workbook.Parent.ActiveWindow
                active_window.SplitColumn = (first_col + last_col) // 2
                active_window.SplitRow = 0
            workbook.Save()
            # workbook.Close(True)
            # excel.Quit()
        return

    # Handle HTML output mode
    if output_format == "html":
        html_table = generate_html_table(
            df_display,
            group_col,
            special_col,
            col_widths,
            group_colors,
            special_colors,
            style_mode,
            persist_special_color,
            nan_placeholder,
            border_style,
            html_max_height,
            synchronized_panels  # Pass the new parameter
        )
        
        try:
            from IPython.display import HTML, display
            # For Jupyter/IPython environments
            display(HTML(html_table))
        except:
            # Fallback for non-Jupyter environments
            
            # Create a temporary HTML file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.html') as f:
                f.write(html_table.encode('utf-8'))
                temp_path = f.name
            
            # Open in browser
            webbrowser.open('file://' + os.path.abspath(temp_path))
            print(f"HTML table saved to: {temp_path}")
            
        return

    # Terminal output mode from here on
    # Get border characters if needed
    if border_style != "none":
        border_chars = {
            "ascii": {"vertical": "|"},
            "unicode": {"vertical": "│"},
            "unicode_heavy": {"vertical": "┃"}
        }
        border_char = border_chars.get(border_style, {"vertical": "|"})["vertical"]

    # Draw top border if needed
    if border_style != "none":
        print(draw_table_border(widths_list, border_style, is_header=None))

    # Print header
    header_parts = []
    for col in df_display.columns:
        col_str = str(col)
        if len(col_str) > col_widths[col]:
            col_str = truncate_text(col_str, col_widths[col])
        header_parts.append(f"{col_str:<{col_widths[col]}}")

    if border_style != "none":
        header = f"{border_char} " + f" {border_char} ".join(header_parts) + f" {border_char}"
    else:
        header = separator.join(header_parts)

    print(header)

    # Draw header separator if using borders
    if border_style != "none":
        print(draw_table_border(widths_list, border_style, is_header=True))

    # Print rows
    for _, row in df_display.iterrows():
        group_val = row[group_col]

        # Get the base color for this row based on its group
        row_color = group_colors[group_val]
        text_color = get_contrasting_color(row_color)

        row_str = ""

        # Start with left border if needed
        if border_style != "none":
            row_str += border_char + " "

        # Process each column separately
        for i, col in enumerate(df_display.columns):
            # Handle NaN values
            val = handle_nan_value(row[col], nan_placeholder)

            # Truncate value if too long
            if len(val) > col_widths[col]:
                val = truncate_text(val, col_widths[col])

            val = val.ljust(col_widths[col])

            # Set default styling based on group color
            bg = row_color if style_mode == "background" else None
            fg = text_color if style_mode == "background" else row_color

            # Apply special coloring if special_col is specified and this column/cell should be special
            if special_col is not None:
                special_val = row[special_col]
                # Make sure NaN is handled properly for color mapping
                if pd.isna(special_val):
                    special_val = np.nan

                special_color = special_colors[special_val]
                special_text_color = get_contrasting_color(special_color)

                is_special = (col == special_col)
                apply_special = is_special or (persist_special_color and col in df_display.columns[
                                                                                df_display.columns.get_loc(
                                                                                    special_col) + 1:])

                # Set coloring based on mode and column
                if apply_special:
                    if style_mode == "text":
                        fg = special_color
                    else:  # background
                        bg = special_color
                        fg = special_text_color

            # Apply style to the column value
            row_str += ansi_style(val, bg=bg, fg=fg)

            # Add separator or border after the column (except the last one)
            if i < len(df_display.columns) - 1:
                if border_style != "none":
                    # Reset style, then add border character with neutral styling
                    row_str += "\033[0m " + border_char + " "
                else:
                    # Reset style, then add separator with neutral styling
                    row_str += "\033[0m" + separator

        # Add right border if needed
        if border_style != "none":
            row_str += "\033[0m " + border_char

        print(row_str)

    # Draw bottom border if needed
    if border_style != "none":
        print(draw_table_border(widths_list, border_style, is_header=False))

    # Print pagination/group info
    print(f"\n{group_info}")

    # Print pagination/grouping guidance
    if group_pagination is not None:
        all_groups = sorted(df_full[group_col].unique())
        print(f"Available groups: {all_groups}")
        print(f"Use print_colored_df(..., group_pagination=['group1', 'group2']) to view specific groups")
    elif page_size is not None:
        print(f"Use print_colored_df(..., current_page=N) to view other pages")


# Modified example function to demonstrate synchronized panels
def example():
    # Create a test dataframe with some NaN values
    df = pd.DataFrame({
        'Group': ['A'] * 6 + ['B'] * 6 + ['C'] * 6,
        'col1': np.random.randint(0, 100, 18),
        'col2': [np.nan if i % 7 == 0 else np.random.randint(0, 100) for i in range(18)],
        'col3': ["This is a very long text that should be truncated" if i % 5 == 0 else i for i in range(18)],
        'special_col': ['X', 'Y', np.nan, 'V', 'Z', 'V', 'X', 'U', 'W', 'Y', 'Z', 'X', 'V', 'Y', 'Z', 'Z', 'Y', 'X'],
        'col4': np.random.randint(0, 100, 18),
        'col5': np.random.randint(0, 100, 18)
    })
    
    # print("\n1. Basic terminal output:")
    # print_colored_df(df, group_col="Group", special_col="special_col", style_mode="background", border_style="unicode")

    # print("\n2. HTML output with standard view:")
    # print_colored_df(df, group_col="Group", special_col="special_col", style_mode="background", 
    #                  border_style="unicode", output_format="html", html_max_height=300)
    
    # print("\n3. HTML output with synchronized panels:")
    # print_colored_df(df, group_col="Group", special_col="special_col", style_mode="background", 
    #                  border_style="unicode", output_format="html", html_max_height=300, 
    #                  synchronized_panels=True)
    
    # Create a large dataframe to demonstrate scrolling
    df_large = pd.DataFrame({
        'Category': np.random.choice(['A', 'B', 'C', 'D'], 100),
        'Value1': np.random.randint(1, 1000, 100),
        'Value2': np.random.randint(1, 1000, 100),
        'Value3': np.random.randint(1, 1000, 100),
        'Status': np.random.choice(['Active', 'Inactive', 'Pending'], 100)
    })
    columns_name = df_large.columns.tolist()
    for _ in range(4):
        df_large = pd.concat([df_large, df_large], ignore_index=True, axis=1)
    df_large.columns = [f"col_({x})" for x in df_large.columns]
    df_large.columns =  columns_name + df_large.columns[len(columns_name):].tolist()
    df_large.sort_values(by=["Category", "Status"], inplace=True)
    df_large['large'] = "1"*50
    df_large['large2'] = "2"*50
    df_large.insert(1, "large", df_large.pop("large"))
    df_large.insert(2, "large2", df_large.pop("large2"))

    # print("\n4. Large HTML table without synchronized panels for better comparison:")
    # print_colored_df(df_large, group_col="Category", special_col="Status", 
    #                  output_format="html", html_max_height=600, synchronized_panels=False)
    
    # print("\n5. Large HTML table with synchronized panels for better comparison:")
    # print_colored_df(df_large, group_col="Category", special_col="Status", 
    #                  output_format="html", html_max_height=600, synchronized_panels=True)

    print("\n6. Large Excel table:")
    print_colored_df(df, group_col="Category", output_name='styled_table', special_col='Status', style_mode='background', output_format="excel",max_column_width=15,persist_special_color=True, synchronized_panels=True)


if __name__ == "__main__":
    example()
